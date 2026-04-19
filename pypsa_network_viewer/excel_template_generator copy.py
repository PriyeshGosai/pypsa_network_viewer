import pypsa
import re
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

_PYPSA_MAJOR = int(pypsa.__version__.split('.')[0])
if _PYPSA_MAJOR < 1:
    raise ImportError(
        f"excel_template_generator requires PyPSA >= 1.0.0, but {pypsa.__version__} is installed"
    )

pypsa.options.api.new_components_api = True

_RESOLUTION_MAP = {
    'H':    timedelta(hours=1),
    '2H':   timedelta(hours=2),
    '4H':   timedelta(hours=4),
    '6H':   timedelta(hours=6),
    '8H':   timedelta(hours=8),
    '0.5H': timedelta(minutes=30),
    '5m':   timedelta(minutes=5),
    '10m':  timedelta(minutes=10),
    '15m':  timedelta(minutes=15),
    '30m':  timedelta(minutes=30),
}

_SNAPSHOT_WEIGHTINGS = {
    'H': 1, '2H': 2, '4H': 4, '6H': 6, '8H': 8,
    '0.5H': 0.5, '5m': 1/12, '10m': 1/6, '15m': 1/4, '30m': 1/2,
}

_TWO_BUS_COMPONENTS = {'lines', 'transformers'}
_BUS_DV_SKIP        = {'global_constraints'}

_INSTRUCTIONS = [
    "PyPSA Configuration Template",
    "",
    "Instructions:",
    "",
    "1. snapshots: Contains the time series data points generated with the specified resolution.",
    "",
    "2. options: Lists all available Input variables per component (bus/carrier fields excluded).",
    "",
    "3. Component sheets (e.g. buses, carriers, generators, loads):",
    "   - Column A: enter component names (row 2 onwards).",
    "   - Carrier column (if any): dropdown linked to the carriers sheet.",
    "   - Bus columns (if any): dropdown linked to the buses sheet.",
    "   - Remaining columns in row 1: dropdown to pick variables from the options list.",
    "     Select a variable, then fill values in the rows below it.",
    "",
    "4. Dynamic variable sheets (e.g. generators-p_max_pu):",
    "   - Column A: snapshot index (1, 2, 3, …).",
    "   - Row 1 (B1 onwards): select component names via dropdown.",
    "   - Fill time-varying values below each selected name.",
    "",
    "5. For links with multiple outputs, bus2/bus3/… and efficiency2/efficiency3/… ",
    "   are included automatically based on the link_outputs setting.",
    "",
    "6. For processes with multiple ports, bus0/bus1/… and rate0/rate1/… ",
    "   are included automatically based on the process_outputs setting.",
    "",
    "7. Hidden sheets can be made visible via Excel's 'Unhide Sheet' option."
]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _validate_params(start_year, start_month, years_duration, months_duration,
                     days_duration, resolution_str, drop_leap_day, link_outputs, process_outputs):
    assert isinstance(start_year, int) and 2024 <= start_year <= 2029, \
        f"start_year must be an int in 2024-2029, got {start_year}"
    assert isinstance(start_month, int) and 1 <= start_month <= 12, \
        f"start_month must be an int in 1-12, got {start_month}"

    if years_duration is not None:
        assert isinstance(years_duration, int) and 1 <= years_duration <= 5, \
            f"years_duration must be None or an int in 1-5, got {years_duration}"
    if months_duration is not None:
        assert isinstance(months_duration, int) and 1 <= months_duration <= 12, \
            f"months_duration must be None or an int in 1-12, got {months_duration}"
    if days_duration is not None:
        assert isinstance(days_duration, int) and 7 <= days_duration <= 31, \
            f"days_duration must be None or an int in 7-31, got {days_duration}"

    dur_count = (years_duration is not None) + (months_duration is not None) + (days_duration is not None)
    assert dur_count == 1, \
        f"Exactly one of years_duration / months_duration / days_duration must be set, got {dur_count} set"

    assert isinstance(resolution_str, str) and resolution_str in _RESOLUTION_MAP, \
        f"resolution_str must be one of {list(_RESOLUTION_MAP.keys())}, got '{resolution_str}'"
    assert isinstance(drop_leap_day, bool), \
        f"drop_leap_day must be a bool, got {type(drop_leap_day)}"
    assert isinstance(link_outputs, int) and link_outputs >= 1, \
        f"link_outputs must be an int >= 1, got {link_outputs}"
    assert isinstance(process_outputs, int) and process_outputs >= 1, \
        f"process_outputs must be an int >= 1, got {process_outputs}"


def _compute_datetime_range(start_year, start_month, years_duration,
                            months_duration, days_duration):
    start_datetime = datetime(start_year, start_month, 1, 0, 0)

    if years_duration is not None:
        end_datetime = datetime(start_year + years_duration, start_month, 1, 0, 0) - timedelta(hours=1)
    elif months_duration is not None:
        total_months = start_month + months_duration
        end_year  = start_year + (total_months - 1) // 12
        end_month = (total_months - 1) % 12 + 1
        end_datetime = datetime(end_year, end_month, 1, 0, 0) - timedelta(hours=1)
    else:
        end_datetime = datetime(start_year, start_month, days_duration, 23, 0)

    return start_datetime, end_datetime


def _generate_timestamps(start_datetime, end_datetime, resolution_delta, drop_leap_day):
    timestamps = []
    current = start_datetime
    while current <= end_datetime:
        if drop_leap_day and current.month == 2 and current.day == 29:
            current += resolution_delta
            continue
        timestamps.append(current)
        current += resolution_delta
    return timestamps


def _clean_version(raw):
    m = re.match(r'(\d+\.\d+\.\d+)', raw)
    return m.group(1) if m else raw


def _hard_bus_fields(comp_name, link_outputs, process_outputs):
    if comp_name == 'links':
        return {'name', 'carrier'} | {f'bus{i}' for i in range(link_outputs + 1)}
    if comp_name == 'processes':
        return {'name', 'carrier'} | {f'bus{i}' for i in range(process_outputs + 1)}
    if comp_name in _TWO_BUS_COMPONENTS:
        return {'name', 'carrier', 'bus0', 'bus1'}
    return {'name', 'carrier', 'bus'} if comp_name not in _BUS_DV_SKIP else {'name', 'carrier'}


def _collect_component_data(link_outputs, process_outputs):
    n = pypsa.Network()
    component_data = {}

    for comp_name in n.components.keys():
        comp     = n.components[comp_name]
        defaults = comp.defaults

        input_vars         = defaults[defaults['status'].str.contains('Input', na=False)].index.tolist()
        dynamic_keys       = set(comp.dynamic.keys())
        dynamic_input_vars = [v for v in input_vars if v in dynamic_keys]

        hard = _hard_bus_fields(comp_name, link_outputs, process_outputs)

        variables, var_types = [], []
        for col in comp.static.columns:
            if col in input_vars and col not in hard:
                variables.append(col)
                var_types.append('static')
        for key in comp.dynamic.keys():
            if key in input_vars and key not in hard:
                variables.append(key)
                var_types.append('dynamic')

        # Add efficiency fields for links with multiple outputs
        if comp_name == 'links' and link_outputs > 1:
            for n_out in range(2, link_outputs + 1):
                variables.append(f'efficiency{n_out}')
                var_types.append('static')

        # Add rate fields for processes with multiple ports
        if comp_name == 'processes' and process_outputs >= 0:
            for n_out in range(0, process_outputs + 1):
                variables.append(f'rate{n_out}')
                var_types.append('static')

        component_data[comp_name] = {
            'input_vars':         input_vars,
            'dynamic_input_vars': dynamic_input_vars,
            'variables':          variables,
            'types':              var_types,
        }

    return n, component_data


def _add_bus_dv(ws, col_idx, start_row=2, end_row=1000):
    dv = DataValidation(type="list",
                        formula1="'buses'!$A$2:$A$1000",
                        allow_blank=False)
    dv.error      = 'Please select a valid bus'
    dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def _add_carrier_dv(ws, col_idx, start_row=2, end_row=1000):
    dv = DataValidation(type="list",
                        formula1="'carriers'!$A$2:$A$1000",
                        allow_blank=True)
    dv.error      = 'Please select a valid carrier'
    dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def _add_var_dv(ws, options_col_letter, options_last_row, first_col, last_col=100):
    dv = DataValidation(type="list",
                        formula1=f"'options'!${options_col_letter}$3:${options_col_letter}${options_last_row}",
                        allow_blank=True)
    dv.error      = 'Please select a valid variable from the options list'
    dv.errorTitle = 'Invalid Variable'
    ws.add_data_validation(dv)
    first_letter = get_column_letter(first_col)
    last_letter  = get_column_letter(last_col)
    dv.add(f"{first_letter}1:{last_letter}1")


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_instructions(wb):
    ws = wb.create_sheet('instructions', 0)
    for row_idx, text in enumerate(_INSTRUCTIONS, 1):
        ws.cell(row=row_idx, column=1, value=text)


def _write_network(wb, pypsa_version):
    ws = wb.create_sheet('network', 1)
    ws['A1'] = 'name'
    ws['B1'] = 'pypsa_version'
    ws['A2'] = 'pypsa_model'
    ws['B2'] = pypsa_version


def _write_snapshots(wb, timestamps, resolution_str):
    ws = wb.create_sheet('snapshots')
    ws['A1'] = 'timestamp'
    ws['B1'] = 'snapshot_weightings'
    weighting = _SNAPSHOT_WEIGHTINGS[resolution_str]
    for row_idx, ts in enumerate(timestamps, 2):
        cell = ws.cell(row=row_idx, column=1, value=ts)
        cell.number_format = 'YYYY-MM-DD HH:MM'
        ws.cell(row=row_idx, column=2, value=weighting)


def _write_options(wb, component_data):
    ws = wb.create_sheet('options')
    options_col_map = {}

    col = 1
    for comp_name, data in component_data.items():
        ws.cell(row=1, column=col,     value=comp_name)
        ws.cell(row=1, column=col + 1, value=comp_name)
        ws.cell(row=2, column=col,     value='Variable')
        ws.cell(row=2, column=col + 1, value='Type')
        for i, (var, vtype) in enumerate(zip(data['variables'], data['types']), 3):
            ws.cell(row=i, column=col,     value=var)
            ws.cell(row=i, column=col + 1, value=vtype)

        last_row = 2 + len(data['variables'])
        options_col_map[comp_name] = (get_column_letter(col), last_row)
        col += 2

    return options_col_map


def _write_component_sheets(wb, n, component_data, options_col_map, link_outputs, process_outputs):
    visible_sheets = {'instructions', 'network', 'snapshots', 'buses', 'carriers', 'generators', 'loads', 'links'}

    # Add processes to visible sheets if it exists in the network
    if 'processes' in n.components.keys():
        visible_sheets.add('processes')

    for comp_name in n.components.keys():
        ws         = wb.create_sheet(comp_name)
        input_vars = component_data[comp_name]['input_vars']

        ws['A1'] = 'Name'
        col = 2

        # Add carrier column if carrier is in input_vars
        if 'carrier' in input_vars:
            ws.cell(row=1, column=col, value='carrier')
            _add_carrier_dv(ws, col)
            col += 1

        if comp_name == 'links':
            for bus_idx in range(link_outputs + 1):
                ws.cell(row=1, column=col, value=f'bus{bus_idx}')
                _add_bus_dv(ws, col)
                col += 1

        elif comp_name == 'processes':
            for bus_idx in range(process_outputs + 1):
                ws.cell(row=1, column=col, value=f'bus{bus_idx}')
                _add_bus_dv(ws, col)
                col += 1

        elif comp_name in _TWO_BUS_COMPONENTS:
            for bus_idx in (0, 1):
                ws.cell(row=1, column=col, value=f'bus{bus_idx}')
                _add_bus_dv(ws, col)
                col += 1

        elif 'bus' in input_vars and comp_name not in _BUS_DV_SKIP:
            ws.cell(row=1, column=col, value='bus')
            _add_bus_dv(ws, col)
            col += 1

        opts_col_letter, opts_last_row = options_col_map[comp_name]
        _add_var_dv(ws, opts_col_letter, opts_last_row, first_col=col)

        if comp_name not in visible_sheets:
            ws.sheet_state = 'hidden'


def _write_dynamic_sheets(wb, n, component_data, timestamps):
    for comp_name in n.components.keys():
        for dyn_var in component_data[comp_name]['dynamic_input_vars']:
            sheet_name = f'{comp_name}-{dyn_var}'[:31]

            ws = wb.create_sheet(sheet_name)
            ws['A1'] = 'snapshot'

            dv = DataValidation(type="list",
                                formula1=f"'{comp_name}'!$A$2:$A$1000",
                                allow_blank=True)
            dv.error      = 'Please select a valid name from the component sheet'
            dv.errorTitle = 'Invalid Entry'
            ws.add_data_validation(dv)
            dv.add("B1:CV1")

            for idx in range(1, len(timestamps) + 1):
                ws.cell(row=idx + 1, column=1, value=idx)

            ws.sheet_state = 'hidden'


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_template(
    output_name,
    start_year=2024,
    start_month=1,
    years_duration=1,
    months_duration=None,
    days_duration=None,
    resolution_str='H',
    drop_leap_day=True,
    link_outputs=1,
    process_outputs=2,
):
    """Generate a PyPSA Excel template workbook.

    Exactly one of years_duration / months_duration / days_duration must be
    set (the other two must be None).

    Parameters
    ----------
    output_name : str
        Path for the output .xlsx file.
    start_year : int
        Start year (2024-2029).
    start_month : int
        Start month (1-12).
    years_duration : int or None
        Duration in years (1-5).
    months_duration : int or None
        Duration in months (1-12).
    days_duration : int or None
        Duration in days (7-31).
    resolution_str : str
        Snapshot resolution: 'H', '2H', '4H', '6H', '8H', '0.5H', '5m', '10m', '15m', '30m'.
    drop_leap_day : bool
        If True, Feb 29 timestamps are skipped.
    link_outputs : int
        Number of link outputs (>=1). Links get bus0 … bus_{link_outputs}.
    process_outputs : int
        Number of process outputs (>=1). Processes get bus0 … bus_{process_outputs}.
    """
    # ── validate ──────────────────────────────────────────────────────────
    _validate_params(start_year, start_month, years_duration, months_duration,
                     days_duration, resolution_str, drop_leap_day, link_outputs, process_outputs)
    print("All parameter checks passed.")

    # ── datetime range & timestamps ───────────────────────────────────────
    start_dt, end_dt = _compute_datetime_range(
        start_year, start_month, years_duration, months_duration, days_duration)

    resolution_delta = _RESOLUTION_MAP[resolution_str]
    timestamps = _generate_timestamps(start_dt, end_dt, resolution_delta, drop_leap_day)
    assert len(timestamps) > 0, "No timestamps generated – check date / duration / resolution settings"

    print(f"Start      : {start_dt.strftime('%d/%m/%Y %H:%M')}")
    print(f"End        : {end_dt.strftime('%d/%m/%Y %H:%M')}")
    print(f"Resolution : {resolution_str}")
    print(f"Leap-day   : {'dropped' if drop_leap_day else 'kept'}")
    print(f"Link outputs: {link_outputs}  ->  bus0 … bus{link_outputs}")
    print(f"Process outputs: {process_outputs}  ->  bus0 … bus{process_outputs}")
    print(f"Snapshots  : {len(timestamps)}")

    # ── version ───────────────────────────────────────────────────────────
    pypsa_version = _clean_version(pypsa.__version__)
    print(f"PyPSA version: {pypsa.__version__}  ->  {pypsa_version}")

    # ── component metadata ────────────────────────────────────────────────
    n, component_data = _collect_component_data(link_outputs, process_outputs)

    # Check if processes component exists
    has_processes = 'processes' in n.components.keys()
    print(f"Process component: {'available' if has_processes else 'not available (requires PyPSA >= 1.1.0)'}")

    # ── build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    _write_instructions(wb)
    _write_network(wb, pypsa_version)
    _write_snapshots(wb, timestamps, resolution_str)
    options_col_map = _write_options(wb, component_data)
    _write_component_sheets(wb, n, component_data, options_col_map, link_outputs, process_outputs)
    _write_dynamic_sheets(wb, n, component_data, timestamps)

    # ── save ──────────────────────────────────────────────────────────────
    wb.save(output_name)

    dyn_count = sum(len(d['dynamic_input_vars']) for d in component_data.values())
    print(f"\nSaved '{output_name}'")
    print(f"  Component sheets   : {len(component_data)}")
    print(f"  Dynamic-var sheets : {dyn_count}")
